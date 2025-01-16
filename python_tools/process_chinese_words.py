import os
import re
import xlsxwriter
import openpyxl

def find_chinese_words(root_path='.'):
    results = []
    pattern = re.compile(r'[\u4e00-\u9fff]+')
    for root, _, files in os.walk(root_path):
        for filename in files:
            filepath = os.path.join(root, filename)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    for lineno, line in enumerate(f, start=1):
                        for match in pattern.findall(line):
                            # print(match, filepath, lineno)
                            results.append((match, filepath, lineno))
            except:
                pass
    return results



if __name__ == "__main__":
    chinese_words = find_chinese_words()
    results = {}
    for match, path, line in chinese_words:
        if match not in results:
            results[match] = []
        results[match].append((path, line))
    # print(results)
    # with xlsxwriter.Workbook('chinese_words.xlsx') as workbook:
    #     worksheet = workbook.add_worksheet()
    #     worksheet.write(0, 0, "Word")       
    #     row = 1

    #     for key in results.keys():         
    #         worksheet.write(row, 0, key)            
    #         row += 1
    workbook = openpyxl.load_workbook('chinese_words.xlsx')
    worksheet = workbook.active

    ch_dict = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            ch_dict[row[0]] = row[1]

    workbook.close()
    # print(ch_dict)

    for ch in chinese_words:
        match, p, ln = ch
        if match in ch_dict:
            if ".\\venv\\" in p:
                continue
            print(match, p, ln)
            with open(p, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            if ln <= len(lines):
                lines[ln - 1] = lines[ln - 1].replace(match, ch_dict[match])
                with open(p, 'w', encoding='utf-8') as f:
                    f.writelines(lines)
    